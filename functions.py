from openpyxl import load_workbook


def get_students_data(excel_file):
    wb = load_workbook(excel_file, read_only=True)
    ws = wb.active
    data = []
    titles = [cell.value for cell in ws[1][4:]]
    subtitles = [cell.value for cell in ws[3][4:]]
    for row in ws.iter_rows(min_row=4, values_only=True, max_col=5 + 18 + 2):
        row_data = {
            "number": row[1],
            "name": row[2],
            "grade": row[3],
        }
        if all(row[1:4]):
            row_data['marks'] = {f"{titles[i - 4] or titles[i - 5]}_{subtitles[i - 4]}": row[i] for i in range(5, 5 + 18)}
            row_data['notes'] = row[5 + 18]
            row_data['picture'] = row[5 + 18 + 1]
            data.append(row_data)
        else:
            break
    return data
